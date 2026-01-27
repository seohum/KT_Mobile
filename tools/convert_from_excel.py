# 직원용 유선정책 엑셀 -> JSON 변환기
# 사용:
#   python tools/convert_from_excel.py "유선정책 단가도해 함수.xlsx"
# 결과:
#   data/wired_policy.json 생성/갱신
import sys, json, datetime
from pathlib import Path
import openpyxl

def extract(wb):
    def extract_wired():
        ws = wb["유선정책표(자동입력)"]
        rows=[]
        r=3
        cur=None
        while r<=ws.max_row:
            a=ws.cell(r,1).value
            b=ws.cell(r,2).value
            c=ws.cell(r,3).value
            if (a in (None,"") and b in (None,"") and c in (None,"")):
                if all(all(ws.cell(rr,cc).value in (None,"") for cc in range(1,4)) for rr in range(r, min(r+5, ws.max_row)+1)):
                    break
                r+=1
                continue
            if a not in (None,""): cur=str(a).strip()
            rows.append({
                "group": cur,
                "internet": None if b in (None,"") else str(b).strip(),
                "tv": None if c in (None,"") else str(c).strip(),
                "base_policy": ws.cell(r,4).value,
                "bundle_policy": ws.cell(r,5).value,
                "total_no_gift": ws.cell(r,6).value,
                "onestop_gift": ws.cell(r,7).value,
                "total_with_onestop": ws.cell(r,8).value,
                "genie3_add": ws.cell(r,9).value,
                "total_with_genie3": ws.cell(r,10).value,
            })
            r+=1
        return rows

    def extract_uit():
        ws = wb["UIT(자동입력)"]
        rows=[]
        r=3
        cur=None
        cur_tier=None
        while r<=ws.max_row:
            a=ws.cell(r,1).value
            b=ws.cell(r,2).value
            c=ws.cell(r,3).value
            d=ws.cell(r,4).value
            if (a in (None,"") and b in (None,"") and c in (None,"") and d in (None,"")):
                if all(all(ws.cell(rr,cc).value in (None,"") for cc in range(1,5)) for rr in range(r, min(r+5, ws.max_row)+1)):
                    break
                r+=1
                continue
            if a not in (None,""): cur=str(a).strip()
            if d not in (None,""): cur_tier=str(d).strip()
            rows.append({
                "group": cur,
                "internet": None if b in (None,"") else str(b).strip(),
                "tv": None if c in (None,"") else str(c).strip(),
                "mobile_tier": cur_tier,
                "u_policy": ws.cell(r,5).value,
                "base_policy": ws.cell(r,6).value,
                "bundle_policy": ws.cell(r,7).value,
                "total_no_gift": ws.cell(r,8).value,
                "onestop_gift": ws.cell(r,9).value,
                "total_with_onestop": ws.cell(r,10).value,
                "genie3_add": ws.cell(r,11).value,
                "total_with_genie3": ws.cell(r,12).value,
            })
            r+=1
        return rows

    return extract_wired(), extract_uit()

def main():
    if len(sys.argv)<2:
        print("엑셀 파일 경로를 넣으세요.")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    wired, uit = extract(wb)
    meta = {
        "source_file": Path(xlsx_path).name,
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "currency_unit": "원",
        "ui_label_rules": "외부 노출 단어(공시/지원금) 미사용. 정책금액/혜택금액으로 표기."
    }
    out = {"meta": meta, "wired_table": wired, "uit_table": uit}
    out_path = Path(__file__).resolve().parents[1] / "data" / "wired_policy.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print("OK:", out_path)

if __name__=="__main__":
    main()
